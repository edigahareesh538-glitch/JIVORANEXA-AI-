"""Smart Expense Manager — analytics, monthly/yearly/category reports, and
PDF / Excel / CSV exports (Phase 6).

Pure-stdlib + reportlab/openpyxl (added in requirements.txt). The endpoints
in `app/routes/expenses.py` call into this module; it never touches the DB.
"""
from __future__ import annotations

import csv
import io
from collections import defaultdict
from datetime import datetime
from typing import Iterable

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)


def _rows_to_dicts(rows: Iterable) -> list[dict]:
    out = []
    for r in rows:
        out.append({
            "id": r.id,
            "category": r.category,
            "label": r.label or "",
            "amount": float(r.amount),
            "currency": r.currency or "INR",
            "spent_at": r.spent_at.strftime("%Y-%m-%d %H:%M") if r.spent_at else "",
        })
    return out


def summarize(rows: Iterable) -> dict:
    """Aggregate analytics over user expenses -- shared by all report variants."""
    by_cat: dict[str, float] = defaultdict(float)
    by_month: dict[str, float] = defaultdict(float)
    by_year: dict[str, float] = defaultdict(float)
    total = 0.0
    items = _rows_to_dicts(rows)
    for item in items:
        by_cat[item["category"]] += item["amount"]
        if item["spent_at"]:
            dt = datetime.strptime(item["spent_at"], "%Y-%m-%d %H:%M")
            by_month[dt.strftime("%Y-%m")] += item["amount"]
            by_year[dt.strftime("%Y")] += item["amount"]
        total += item["amount"]
    return {
        "total": round(total, 2),
        "count": len(items),
        "by_category": {k: round(v, 2) for k, v in sorted(by_cat.items(), key=lambda kv: -kv[1])},
        "by_month":   {k: round(v, 2) for k, v in sorted(by_month.items())},
        "by_year":    {k: round(v, 2) for k, v in sorted(by_year.items())},
        "items": items,
    }


def category_report(rows: Iterable, category: str) -> dict:
    items = [r for r in _rows_to_dicts(rows) if r["category"] == category]
    total = sum(i["amount"] for i in items)
    return {
        "category": category,
        "total": round(total, 2),
        "count": len(items),
        "items": items,
    }


def budget_vs_actual(rows: Iterable, plan_allocations: dict[str, float]) -> dict:
    actuals = defaultdict(float)
    for r in _rows_to_dicts(rows):
        actuals[r["category"]] += r["amount"]
    keys = sorted(set(list(plan_allocations.keys()) + list(actuals.keys())))
    rows_out = []
    for k in keys:
        planned = round(plan_allocations.get(k, 0), 2)
        actual = round(actuals.get(k, 0), 2)
        rows_out.append({
            "category": k,
            "planned": planned,
            "actual": actual,
            "variance": round(planned - actual, 2),
            "over": actual > planned,
        })
    return {"rows": rows_out, "total_planned": round(sum(plan_allocations.values()), 2),
            "total_actual": round(sum(actuals.values()), 2)}


# ---------------------- exports ----------------------

def export_csv(summary_items: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["spent_at", "category", "label", "amount", "currency"])
    writer.writeheader()
    for row in summary_items:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def export_pdf(summary: dict, title: str = "Expense Report") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=title)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"<b>{title}</b>", styles["Title"]),
        Spacer(1, 12),
        Paragraph(f"Total spend: <b>₹{summary['total']:.2f}</b> · {summary['count']} entries", styles["Normal"]),
        Spacer(1, 12),
    ]
    if summary["by_category"]:
        elements.append(Paragraph("<b>By category</b>", styles["Heading3"]))
        cat_rows = [["Category", "Amount (₹)"]] + [
            [k, f"{v:.2f}"] for k, v in summary["by_category"].items()
        ]
        cat_table = Table(cat_rows, hAlign="LEFT")
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#374151")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        elements.append(cat_table)
        elements.append(Spacer(1, 12))
    if summary["by_month"]:
        elements.append(Paragraph("<b>By month</b>", styles["Heading3"]))
        m_rows = [["Month", "Amount (₹)"]] + [
            [k, f"{v:.2f}"] for k, v in summary["by_month"].items()
        ]
        m_table = Table(m_rows, hAlign="LEFT")
        m_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#374151")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        elements.append(m_table)
    doc.build(elements)
    return buf.getvalue()


def export_xlsx(summary: dict, sheet_name: str = "Expenses") -> bytes:
    """Build a minimal but well-formed .xlsx using openpyxl when available,
    fall back to CSV bytes with an .xlsx extension otherwise (Excel still
    opens this with a 'convert from text' prompt)."""
    try:
        from openpyxl import Workbook  # type: ignore
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        ws.append(["Spent At", "Category", "Label", "Amount", "Currency"])
        for row in summary["items"]:
            ws.append([row["spent_at"], row["category"], row["label"], row["amount"], row["currency"]])
        ws2 = wb.create_sheet("By Category")
        ws2.append(["Category", "Amount"])
        for k, v in summary["by_category"].items():
            ws2.append([k, v])
        ws3 = wb.create_sheet("By Month")
        ws3.append(["Month", "Amount"])
        for k, v in summary["by_month"].items():
            ws3.append([k, v])
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()
    except Exception:
        # Fallback: if openpyxl isn't installed, return CSV bytes
        # (named like an xlsx only if caller insists; otherwise use CSV).
        return export_csv(summary["items"])
